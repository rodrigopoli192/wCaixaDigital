"""
Relatorios services - Export services for reports.
"""

import io
from datetime import datetime
from decimal import Decimal

from django.http import HttpResponse
from django.template.loader import render_to_string

try:
    from weasyprint import HTML

    HAS_WEASYPRINT = True
except ImportError:
    HAS_WEASYPRINT = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ExportService:
    """Service for exporting reports to PDF and XLSX."""

    @staticmethod
    def to_pdf(
        title: str,
        columns: list[dict],
        rows: list[dict],
        totals: dict | None = None,
        filters: dict | None = None,
        tenant_name: str = "",
    ) -> HttpResponse:
        """
        Generate a PDF report.

        Args:
            title: Report title
            columns: List of column definitions [{"key": "field", "label": "Label", "align": "left|right"}]
            rows: List of row data dicts
            totals: Optional totals dict {"field": value}
            filters: Optional applied filters dict
            tenant_name: Tenant name for header
        """
        if not HAS_WEASYPRINT:
            return HttpResponse("WeasyPrint não está instalado.", status=500)

        context = {
            "title": title,
            "columns": columns,
            "rows": rows,
            "totals": totals or {},
            "filters": filters or {},
            "tenant_name": tenant_name,
            "generated_at": datetime.now(),
        }

        html_string = render_to_string("relatorios/pdf/base_pdf.html", context)
        html = HTML(string=html_string)
        pdf_file = html.write_pdf()

        response = HttpResponse(pdf_file, content_type="application/pdf")
        filename = f"{title.lower().replace(' ', '_')}_{datetime.now():%Y%m%d_%H%M%S}.pdf"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    @staticmethod
    def to_xlsx(
        title: str,
        columns: list[dict],
        rows: list[dict],
        totals: dict | None = None,
        filters: dict | None = None,
    ) -> HttpResponse:
        """
        Generate an XLSX report.

        Args:
            title: Report title (used for sheet name and filename)
            columns: List of column definitions [{"key": "field", "label": "Label"}]
            rows: List of row data dicts
            totals: Optional totals dict {"field": value}
            filters: Optional applied filters dict
        """
        if not HAS_OPENPYXL:
            return HttpResponse("openpyxl não está instalado.", status=500)

        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel sheet name limit

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        total_font = Font(bold=True, size=11)
        total_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

        border = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")

        # Filters row (if any)
        current_row = 2
        if filters:
            filter_text = " | ".join(f"{k}: {v}" for k, v in filters.items() if v)
            if filter_text:
                ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
                ws.cell(row=2, column=1, value=f"Filtros: {filter_text}")
                current_row = 3

        # Empty row
        current_row += 1

        # Header row
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=col["label"])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        current_row += 1

        # Data rows
        for row_data in rows:
            for col_idx, col in enumerate(columns, 1):
                value = row_data.get(col["key"], "")
                if isinstance(value, Decimal):
                    value = float(value)
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.border = border
                if col.get("align") == "right":
                    cell.alignment = Alignment(horizontal="right")
            current_row += 1

        # Totals row
        if totals:
            for col_idx, col in enumerate(columns, 1):
                value = totals.get(col["key"], "TOTAL" if col_idx == 1 else "")
                if isinstance(value, Decimal):
                    value = float(value)
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.font = total_font
                cell.fill = total_fill
                cell.border = border
                if col.get("align") == "right":
                    cell.alignment = Alignment(horizontal="right")

        # Auto-adjust column widths
        for col_idx, col in enumerate(columns, 1):
            max_length = len(col["label"])
            for row in ws.iter_rows(
                min_row=4, max_row=current_row, min_col=col_idx, max_col=col_idx
            ):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)

        # Generate response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        filename = f"{title.lower().replace(' ', '_')}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


def format_currency(value: Decimal | float | None) -> str:
    """Format a value as Brazilian currency."""
    if value is None:
        return "R$ 0,00"
    return f"R$ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
